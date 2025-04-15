from django.shortcuts import render,redirect
from django.core.paginator import Paginator
from openpyxl import Workbook
from django.http import HttpResponse
from .models import Author,Book,BorrowRecord


# Create your views here.

def home(request):

    return render(request,'home.html')


def index(request):

    if request.method == 'POST':
        name = request.POST.get('name')
        email = request.POST.get('email')
        bio = request.POST.get('bio')

        
        Author.objects.create(name=name, email=email, bio=bio)
        return redirect('index')
    
    
    return render(request,'index.html') 

def add_book(request):


    if request.method == "POST":

        title = request.POST.get('title')
        genre = request.POST.get('genre')
        published_date = request.POST.get('published_date')
        author_id=request.POST.get('author') 

    
           
        author = Author.objects.get(id=author_id)

        
            
        if author:

             Book.objects.create(
                title=title,
                genre=genre,
                published_date=published_date,
                 author=author
           ) 

             return redirect('home')
        else:
           
            return render(request, 'add_book.html', {'error': 'Author not found!'})

    authors = Author.objects.all()
    return render(request, 'add_book.html', {'authors': authors}) 

def show_books(request):
    
    all_books = Book.objects.all()
    paginator = Paginator(all_books, 5) 
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'show_books.html', {'page_obj': page_obj})


def borrow_record(request):
    if request.method == 'POST':
       if request.method == 'POST':
        user_name = request.POST['user_name']
        borrow_date = request.POST['borrow_date']
        return_date = request.POST['return_date']
        book = Book.objects.get(id=request.POST['book'])

        BorrowRecord.objects.create(
            user_name=user_name,
            borrow_date=borrow_date,
            return_date=return_date,
            book=book
        )
        return redirect('home') 
    
    books = Book.objects.all()
    return render(request, 'borrow_record.html', {'books': books})

from .models import Author  

def book_list(request):
    authors = Author.objects.all()
    paginator = Paginator(authors, 3)  

    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'book_list.html', {'page_obj': page_obj}) 

def export_authors_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Authors"

    ws.append(['ID', 'Name'])  
    for author in Author.objects.all():
        ws.append([author.id, author.name])  

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="authors.xlsx"'
    wb.save(response)
    return response
def export_books_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Books"

    ws.append(['ID', 'Title', 'Author', 'Published Date'])  

    for book in Book.objects.all():
        ws.append([book.id, book.title, book.author.name, book.published_date])  

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="books.xlsx"'
    wb.save(response)
    return response
def export_borrow_records_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Borrow Records"

    ws.append(['ID', 'User Name', 'Book Title', 'Borrow Date', 'Return Date'])  

    for record in BorrowRecord.objects.all():
        ws.append([record.id, record.user_name, record.book.title, record.borrow_date, record.return_date])  

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="borrow_records.xlsx"'
    wb.save(response)
    return response


def export(request):
    return render(request,'export.html')